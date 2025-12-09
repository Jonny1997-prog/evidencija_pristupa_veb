import os
from openpyxl import load_workbook

from app import init_db, get_db


def import_lookups_from_excel(xlsx_path: str) -> None:
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel fajl nije pronađen: {xlsx_path}")

    init_db()

    wb = load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb["Baza za padajuci meni"]
    except KeyError:
        raise KeyError("Sheet 'Baza za padajuci meni' nije pronađen u Excel fajlu.")

    employees = set()
    objects = set()
    destinations = set()


    for row in ws.iter_rows(min_row=2):
        val_emp = row[0].value
        val_obj = row[2].value
        val_dest = row[4].value

        if isinstance(val_emp, str) and val_emp.strip():
            employees.add(val_emp.strip())

        if isinstance(val_obj, str) and val_obj.strip():
            objects.add(val_obj.strip())

        if isinstance(val_dest, str) and val_dest.strip():
            destinations.add(val_dest.strip())

    conn = get_db()
    cur = conn.cursor()


    cur.execute("DELETE FROM lookups WHERE type IN ('employee', 'object', 'destination')")

    cur.executemany(
        "INSERT INTO lookups (type, value) VALUES ('employee', ?)",
        [(e,) for e in sorted(employees)],
    )
    cur.executemany(
        "INSERT INTO lookups (type, value) VALUES ('object', ?)",
        [(o,) for o in sorted(objects)],
    )
    cur.executemany(
        "INSERT INTO lookups (type, value) VALUES ('destination', ?)",
        [(d,) for d in sorted(destinations)],
    )

    conn.commit()
    conn.close()


if __name__ == "__main__":
    default_path = os.path.join(os.path.dirname(__file__), "VEB APP.xlsx")
    import_lookups_from_excel(default_path)
    print("Lookup vrednosti uspešno uvezene iz Excel-a.")
